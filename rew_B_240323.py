# load bibliothek
import pandas as pd
from load_data import link
from load_data import adres
from checker import numer, head
from data_pandas import data_pandas

# import def for page
from format_openpyxl import set_page_settings
# import def for cell
from format_openpyxl import apply_format

# change number on character
from openpyxl.utils.cell import get_column_interval
# for openpyxl make border and side(?)
from openpyxl.styles.borders import Border, Side
# make possible change size of test(font), and roatation(alignment), and fill
from openpyxl.styles import Font, Alignment, PatternFill

# Preaper link
link = link()
# create df
df = adres(link)
# check numer
numer(df)
# check headers
head(df)
# convert df in pandas
df = data_pandas(df)

# sheet name zmienna dla nazwy excela
sheet_name = ("test" + ".xlsx")

# with pd.excelwriter - open excel, realise code below, close excel
# in bracket #1 = name od doc, when we saved channed, #2 engine, #writer
# - save?)
with pd.ExcelWriter(sheet_name, engine="openpyxl") as writer:
    # import df
    df.to_excel(writer, sheet_name, startrow=0, startcol=0)
    # please where code must be realised
    book = writer.book
    # please where code must be realised (sheet)
    ws = writer.sheets[sheet_name]

    # create def who adds parameters for page
    # font settings
    font = Font(name='Arial', size=5.5, bold=True)
    # alignment settings for cell
    alignment = Alignment(horizontal='center', vertical='center')
    # alignment settings for headers
    alignment_headers = Alignment(horizontal='center', vertical='center',
                                  textRotation=180)
    # fill for part
    patern_fill_part = PatternFill()
    # fill for assembilies
    patern_fill_assemblies = PatternFill(start_color='ADD8E6',
                                         end_color='ADD8E6', fill_type='solid')
    # seeting for border
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    # seeting for def
    set_page_settings(ws, '1160', 'autor MD')

    # show last row and last colin df
    nrows, ncol = df.shape
    # start col
    start_column = 1
    # end col
    end_column = ncol+1
    # col_char start and end
    col_char = get_column_interval(start_column, end_column)
    # range of list charcter ,py convert A to 65 and Z to 90
    col_char_range = col_char[0:ncol+1]
    # max_list
    max_list = []
    # max_list for col
    max_list_x = []
    # this loop create list of max char in col
    for x in col_char_range:
        # max = = 0 size of cell
        max = 0
        # this loop set max size of rows
        # +2 beacoue need 1 extra, and 2 extra for headers (without first line)
        for y in range(2, nrows+2):
            # coridnate of cell
            cell = f'{x}{y}'
            # value of coridnate of cell
            cell_value = ws[cell].value
            # show len of cell, +1 becouse some time col is empty
            cell_size = len(str(cell_value))+1
            # check what is bigger
            if cell_size > max:
                # name of col
                max_x = x
                # set up max length
                max = cell_size
            # if cell is not bigger pass
            else:
                pass
        # ws.column_dimensions[x].width = max * 0.5 + 1.4
        max_list.append([max, max_x])
    # loop for columns, find where col = "indeks czesci"
    for x in col_char_range:
        # loop for rows
        for y in range(1, nrows + 2):
            # coridnate of cell
            cell = f'{x}{y}'
            # coridnate of cell value
            cell_value = ws[cell].value
            # find where is col = "Indeks Czesci":
            if cell_value == "Indeks Czesci":
                # take first character 'A' ,'B' or other
                indeks_czesci = cell[0]
    # assemblies list creator
    assemblies_list = []
    # loop for columns assemblies list creator
    for x in col_char_range:
        # for col = indeks_czesci
        if x == indeks_czesci:
            # loop for rows
            for y in range(1, nrows + 2):
                # coridnate of cell
                cell = f'{x}{y}'
                # return value of cell
                cell_value = ws[cell].value
                # find assembiles
                if cell_value[0] == "Z":
                    assemblies_list.append(y)
                # if cell is not bigger pass
                else:
                    pass
    # list creator
    # create empty list of headers
    headers_list = []
    # loop for col hedaers list creator
    for x in col_char_range:
        # headers row = 1
        y = 1
        # coridnate of cell
        cell = f'{x}{y}'
        # create list headers cell
        headers_list.append(cell)

    # match size of col
    for x in max_list:
        # x[1] = number of col, x[0]= max size, 0,5 +1,4 parameter
        # set up size of col for max char in cell
        ws.column_dimensions[(x[1])].width = (x[0]) * 0.45 + 1.4

    # match setting for headers
    for cell in headers_list:
        # rotation text
        # ws[cell].alignment = style_alignment
        apply_format(ws, cell, ws[cell].value, font, alignment_headers,
                     patern_fill_part, thin_border)
    # matchsetting for whole excel
    for x in col_char_range:
        for y in range(2, nrows + 2):
            cell = f'{x}{y}'
            apply_format(ws, cell, ws[cell].value, font, alignment,
                         patern_fill_part, thin_border)
    # matchsetting for assemblies
    for y in assemblies_list:
        # loop for col
        for x in col_char_range:
            # cell cord
            cell = f'{x}{y}'
            # fill cell
            apply_format(ws, cell, ws[cell].value, font, alignment,
                         patern_fill_assemblies, thin_border)
print(f'openpyxl finished sucessfull check file --> {sheet_name}')
