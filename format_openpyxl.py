# first def
# add margin (down, top and another)
from openpyxl.worksheet.page import PageMargins
# f or openpyxl make data
import datetime


def set_page_settings(ws, project, author):
    # create margins left and right
    ws.page_margins = PageMargins(left=0.5, right=0.4)
    # create header with random name
    ws.oddHeader.center.text = project
    # size of header character
    ws.oddHeader.center.size = 14
    # tyoe if font
    ws.oddHeader.center.font = "Tahoma,Bold"
    # color of head
    ws.oddHeader.center.color = "CC3366"
    # lef footer
    ws.oddFooter.left.text = "Strona &P z &N"
    # number of pages in doc
    ws.evenFooter.left.text = "Strona &P z &N"
    # current time
    data = datetime.datetime.now()
    #  add actual time
    ws.append([data])
    # addcurrent time in right footer
    ws.oddFooter.right.text = "&d"
    # add time in every page
    ws.evenFooter.right.text = "&d"
    # add who created data
    ws.oddFooter.center.text = author
    # freeze first row , doesn t work
    ws.freeze_panes = "A2"
    # printer Settings - page orientation
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    # printer Settings - page  size
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def apply_format(ws, cell, value, font, alignment, patern_fill, thin_border):
    ws[cell] = value
    ws[cell].font = font
    ws[cell].alignment = alignment
    ws[cell].fill = patern_fill
    ws[cell].border = thin_border
