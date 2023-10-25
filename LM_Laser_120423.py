###obroka danych w pandasie)
#zaczytanie pandasa
import pandas as pd
#podanie adresu
url='XXX'
url='https://drive.google.com/uc?id=' + url.split('/')[-2]
#przypisanie dataframe i kodowanie wg win 1250 - znaki pl
df = pd.read_csv(url,encoding='windows-1250')
#I z kolumn o nazwach jak nizej wyrzuca spacje
df["Indeks Czesci"] = df["Indeks Czesci"].str.replace(' ', '')

#III A czyszcze material 1
df= df.replace({"Materiał": {"Materiał <nieokreślony>":"X"}})
#III B czyszcze material 2
df= df.replace({"Materiał": {"Element handlowy":"handlowy"}})
#III C czyszcze gabaryty
df= df.replace("Część wieloobiektowa","")

#IV A dopisanie zer, bez tego nie pusci zmianay na floata
df.loc[:,["Grubość ","Dlugosc","Szerokość"]] = df.loc[:,["Grubość ","Dlugosc","Szerokość"]].replace("", "0")
#IV B zamiana kolumn z gabarytami na float (inaczej round nie przejdzie)
df.loc[:,["Grubość ","Dlugosc","Szerokość"]]=df.loc[:,["Grubość ","Dlugosc","Szerokość"]].astype(float)
#IV C zaokraglenie do rządanych wartosci
df=df.round({"Masa":1, "Grubość ":0, "Dlugosc":0, "Szerokość":0})
#IV D Celem jest zamiana 1 na X , a 0 na nic "" w operacjach technologicznych(laser itd)
#IV E przedzial kolumn od giecia do kolumn --> wyrzucenie pustych wartosci (bez tego nei dziala)
df.loc[:,"Gięcie":"Standardowa wypalka"] = df.loc[:,"Gięcie":"Standardowa wypalka"].fillna(0)
#IV F konwersja na int (gubi teoretyczne wartoci dziesietne)
df.loc[:,"Gięcie":"Standardowa wypalka"] = df.loc[:,"Gięcie":"Standardowa wypalka"].astype("int")
#IV G przejscie teraz na str zeby moc zamieniac symbole
df.loc[:,"Gięcie":"Standardowa wypalka"] = df.loc[:,"Gięcie":"Standardowa wypalka"].astype("str")
#IV H zamiania 0 na nic ""
df.loc[:,"Gięcie":"Standardowa wypalka"] = df.loc[:,"Gięcie":"Standardowa wypalka"].replace("0", "")
#IV I zamian 1 na X
df.loc[:,"Gięcie":"Standardowa wypalka"] = df.loc[:,"Gięcie":"Standardowa wypalka"].replace("1", "X")
#IV J zamiana w indeksie spacji na nic ""
df["Indeks Czesci"] = df["Indeks Czesci"].str.replace(' ', '')

#V A zapis do excela opcjonalnie
#df.to_excel("test_poj_excel.xlsx")

print("koniec pandasa")

### VI A MODULY pobrane
# VI B pobiera modul pokazujacy wspolrzedne kolumn i wierszy
### Przejscie do sekcji openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
# VI C inny sposob na pokazanie komorki (zamienia cyfre na litere np 1 = A)
from openpyxl.utils import get_column_letter
# VI Di nny sposob na sposob na pokazanie komorki
import openpyxl.utils.cell
# IX A pobranie wypelnienia
from openpyxl.styles import PatternFill
# X A pobranie polozenia tekstu (np 180 stopni)
from openpyxl.styles import Alignment
# XI A wielkosc tekstu
from openpyxl.styles import Font
###              ramka
from openpyxl.styles.borders import Border, Side
# XV dodanie moduly daty
import datetime

### VII A PRZYGOTOWANIE PLIKU I ARKUSZA
# sheet name zmienna dla nazwy excela
sheet_name = ("MOJA NAZWA" + ".xlsx")
# VIIB pyton ksiazka strona 147
# with pd.excelwriter - otwiera na chwile excela, przerabia kod nizej i zamyka,
# nie trzeba dawad open, write, close(
# zawias ( pliku do ktorego zapisze dane", modul/silnik na jakim bebe pracowal)
# writer - zapisywacz
with pd.ExcelWriter(sheet_name, engine="openpyxl") as writer:
    # VII C eksportuje df=Dataframe
    #  ()nazwa pocodzi z writera, nazwa sheeta, polozenie)
    df.to_excel(writer, sheet_name, startrow=0, startcol=0)
    # VII D okreslenie gdzie maja zachodzi wszyskie polcenia kodu(plik)
    book = writer.book
    # VII E okreslenie gdzie maja zachodzi wszyskie polcenia kodu(arkusz)
    ws = writer.sheets[sheet_name]
    # VII F max wiersz + max kolumna na podstawie powierzchni skrajnych polozen df
    nrows, ncol = df.shape

    # IX B Formatowanie koloru
    style = PatternFill(fgColor="D9D9D9", fill_type="solid")
    # X B  obrocenie tekstu o 180 stopni
    style.Alignment = Alignment(textRotation=180)
    # XI B przypisanie wielkosci tekstu
    fontStyle = Font(size="6")
    # XII A obramowanie
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # XIII A niebieskie zloz
    style_z = PatternFill(start_color="00666699", end_color="009999FF", fill_type="solid")

    ws.freeze_panes = 'A2'
    # ws.row_dimensions[1].font = Font(bold=True)

    # marginexy jak dodam dolny ,gorny ,stopkowy to wywala stopke i naglowek
    from openpyxl.worksheet.page import PageMargins

    ws.page_margins = PageMargins(left=0.5, right=0.4)

    # XIV A dodanie naglowka i formatowanie go
    # XIV B tytul
    ws.oddHeader.center.text = "test"
    # XIV B wielkosc czcionki
    ws.oddHeader.center.size = 14
    # XIV C czcionka
    ws.oddHeader.center.font = "Tahoma,Bold"
    # XIV D kolor
    ws.oddHeader.center.color = "CC3366"
    # XV A Dodanie stopek
    # XV B lewy dol dodanie numer strony / max ilosc stron
    ws.oddFooter.left.text = "Strona &P z &N"
    # XV C lewy dol dodanie numer strony / max ilosc stron
    ws.evenFooter.left.text = "Strona &P z &N"
    # XV D zaczytanie obecnego czasu
    data = datetime.datetime.now()
    # XV E dodanie go do listy (po co?)
    ws.append([data])
    # XV F dodanie daty w prawy rogu
    ws.oddFooter.right.text = "&d"
    # XV G powielenie daty na kazdej stronie
    ws.evenFooter.right.text = "&d"

    # XVI A Printer Settings
    # XVI B polozenie kartki
    ws.page_setup.orientation = 'portrait'
    # XVI C rozmiar kartki
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # XVII A  Set the print titles ### nietesty nie dziala!!!!!!!!!!!!!!1
    # ws.print_title_cols = 'A:B' # the first two cols
    # ws.print_title_rows = '1:1' # the first row

    # VIII A przejscie po kolumnach, nie wiem czemu te +1 ???!!!!
    for x in range(0, ncol + 1):
        # VIII B przypisanie wartosci wstepnej dla szerkosci kolumny
        width_x = [0]
        # VIII zamian cyfry kolumna na litera np 1 = A , 4 = D po co te +1??
        x = openpyxl.utils.cell.get_column_letter(1 + int(x))
        # XIII B przypisanie warunku brzegwo, bez teo nei dziala ????!!!
        nazwy = None
        # VIII D przejscie po wierszach od 0 nie moze zaczac, ale po +2?
        for y in range(1, nrows + 2):
            # VIII E wygenerowanie wspolrzednych x to kolumny, y to wierze
            xy = ((str(x) + str(y)))
            # X C w celu skrocenia pisania przypisanie wartosi cell
            v_xy = ws[str(xy)]

            # X D gdy wiersz = 1 bede obracal tekst o 180 stopni ,nie licze szerokosci
            if y == 1:
                # X E obroc wieresz zgodnei z X B
                v_xy.alignment = Alignment(textRotation=180, horizontal="center", vertical="center")
                # XI C wielkosc czcionki z XI B
                v_xy.font = fontStyle
                # XII B obramowanie
                v_xy.border = thin_border
                # XIII C wylapanie dla ktorej kolumny jest indeksy czesci (niebieskie tlo)
                if v_xy.value == "Indeks Czesci":
                    # XIII D wylapanie dla ktorej komorki indeks jest rowny indeks czesci
                    nazwy = (str(xy)[0])
            # VIII O liczenie wierszy ktore sa rozne od 1 i kolumn
            else:
                # XIII F jak indeks czesci bedzie = forowanej kolumnie (leci tylko po wierszach)
                if nazwy == x:
                    # XIII G nowa zmienna ,bo na starej nei dziala?!
                    z = v_xy.value
                    # VIII L liczenie dlugosci znakow
                    legth = (len(str(ws[str(xy)].value)))
                    # VIII M liczenie dlugosci znakow
                    width_x.append(legth)
                    # XIII H jak pierwszy symbol to Z w kolumnie indekss czesci
                    if z[0] == "Z":
                        # XIII I leci teraz po wierszach, na nowej x2 ,bo stara juz wykorzystana
                        for x2 in range(0, ncol + 1):
                            # XIII J zamian cyfry kolumna na litera np 1 = A , 4 = D po co te +1??
                            x2 = openpyxl.utils.cell.get_column_letter(1 + int(x2))
                            # XIII K wygenerowanie wspolrzednych x2 to kolumny, y to wierze
                            x2y = ((str(x2) + str(y)))
                            # XIII L pozsksanie wartosi cell
                            v_x2y = ws[str(x2y)]
                            # XIII M pomalowanie
                            v_x2y.fill = style_z
                            # XIII N wielkosc czcionki
                            v_x2y.font = fontStyle
                            # XIII O obramowanie
                            v_x2y.border = thin_border
                            # XIII P wyrownanie
                            v_x2y.alignment = Alignment(horizontal="center", vertical="center")
                    # XIII R ustawienia dla kolumny czesc i nie zaczynajaych sie na Z
                    else:
                        # XIII N wielkosc czcionki
                        v_xy.font = fontStyle
                        # XIII O obramowanie
                        v_xy.border = thin_border
                        # XIII P wyrownanie
                        v_xy.alignment = Alignment(horizontal="center", vertical="center")
                # VIII P Przejcscie po reszcie komorek (  przejscie sie dubluje z XIII I)
                else:
                    # VIII G liczenie dlugosci znakow
                    legth = (len(str(ws[str(xy)].value)))
                    # VIII H wygenerowanie listy dla jednej kolumny
                    width_x.append(legth)
                    # XI D wielkosc czcionki z XI B
                    v_xy.font = fontStyle
                    # XIII O obramowanie
                    v_xy.border = thin_border
                    # XIII P wyrownanie
                    v_xy.alignment = Alignment(horizontal="center", vertical="center")

        # XVII A Okreslenie maksymalnego obszaaru drukowania
        # XVII B okreslenie maksymalnej kolumny
        max_col = openpyxl.utils.cell.get_column_letter(1 + int(ncol))
        # XVII C okreslenie masymalnej wsporzlednej xy
        xy_max = ((str(max_col) + str(nrows + 2)))
        # XVII D zmienna druk do wprowadzenie
        druk = f'\"B1:{xy_max}\"'
        # XVII E okreslenie obszaru drukowania
        ws.print_area = f"B1:{xy_max}"

        # VIII I okreslenie najszerszej komorki w kolumnie
        max_legth = ((max(width_x)))
        # VIII J przypisanie najwiekszej szerokosci tekstu dla kolumny
        ws.column_dimensions[x].width = max_legth * 0.5 + 1.4
    print("koniec openpyxl")
